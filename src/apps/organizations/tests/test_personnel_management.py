from __future__ import annotations

import io
from datetime import date

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.organizations.authority_models import OperationalAuthorityGrant
from apps.organizations.models import (
    Division,
    Employee,
    EmployeeOperationalRight,
    OperationalRightDefinition,
    Organization,
    Position,
    Workplace,
)
from apps.organizations.personnel_management_models import (
    EmployeeContactProfile,
    PersonnelChangeAction,
    PersonnelChangeRecord,
    PersonnelImportBatch,
    PersonnelImportStatus,
)
from apps.organizations.personnel_management_services import build_personnel_template
from tests.credential_fixtures import ephemeral_credential


class PersonnelManagementViewTests(TestCase):
    @classmethod
    def setUpTestData(cls) -> None:
        user_model = get_user_model()
        cls.user = user_model.objects.create_user(
            username="personnel-editor",
            password=ephemeral_credential("PersonnelEditor"),
        )
        cls.organization = Organization.objects.create(
            code="PM-ORG",
            name="Синтетическая эксплуатационная организация",
            short_name="PM-ORG",
        )
        cls.division = Division.objects.create(
            organization=cls.organization,
            code="OPS",
            name="Оперативная служба",
        )
        cls.second_division = Division.objects.create(
            organization=cls.organization,
            code="RZA",
            name="Группа РЗА",
        )
        cls.position = Position.objects.create(
            organization=cls.organization,
            code="ELECTRICIAN",
            name="Электромонтёр",
            is_operational=True,
        )
        cls.workplace = Workplace.objects.create(
            organization=cls.organization,
            division=cls.division,
            code="SHIFT",
            name="Сменный персонал",
        )
        cls.right_definition, _ = OperationalRightDefinition.objects.get_or_create(
            code="switching_operation",
            defaults={
                "name": "Производство переключений",
                "category": OperationalRightDefinition.Category.SWITCHING,
                "display_order": 10,
            },
        )

    def setUp(self) -> None:
        self.client.force_login(self.user)

    def employee_payload(self, **overrides):
        payload = {
            "employee-organization": str(self.organization.id),
            "employee-division": str(self.division.id),
            "employee-position": str(self.position.id),
            "employee-workplace": str(self.workplace.id),
            "employee-personnel_number": "PM-001",
            "employee-last_name": "Тестов",
            "employee-first_name": "Иван",
            "employee-middle_name": "Петрович",
            "employee-employment_start": "2026-01-01",
            "employee-employment_end": "",
            "employee-is_active": "on",
            "employee-change_reason": "Заявка на ведение справочника",
            "contact-primary_phone": "+7 900 000-00-01",
            "contact-operational_phone": "1001",
            "contact-email": "operator@example.test",
            "contact-availability_schedule": "круглосуточно",
            "contact-is_round_the_clock": "on",
            "contact-note": "Синтетические данные",
        }
        payload.update(overrides)
        return payload

    def create_employee(self) -> Employee:
        employee = Employee.objects.create(
            organization=self.organization,
            division=self.division,
            position=self.position,
            workplace=self.workplace,
            personnel_number="PM-001",
            last_name="Тестов",
            first_name="Иван",
            middle_name="Петрович",
            employment_start=date(2026, 1, 1),
        )
        EmployeeContactProfile.objects.create(employee=employee)
        return employee

    def test_create_employee_card_with_contacts_and_audit(self) -> None:
        response = self.client.post(
            reverse("organizations:employee_create"),
            self.employee_payload(),
        )

        self.assertEqual(response.status_code, 302)
        employee = Employee.objects.get(
            organization=self.organization,
            personnel_number="PM-001",
        )
        self.assertEqual(employee.contact_profile.operational_phone, "1001")
        self.assertTrue(employee.contact_profile.is_round_the_clock)
        audit = PersonnelChangeRecord.objects.get(employee=employee)
        self.assertEqual(audit.action, PersonnelChangeAction.CREATE)
        self.assertEqual(audit.after_snapshot["full_name"], employee.full_name)

    def test_edit_existing_card_moves_employee_without_duplication(self) -> None:
        employee = self.create_employee()
        payload = self.employee_payload(
            **{
                "employee-division": str(self.second_division.id),
                "employee-change_reason": "Перевод в группу РЗА",
                "contact-operational_phone": "2002",
            }
        )

        response = self.client.post(
            reverse(
                "organizations:employee_edit",
                kwargs={"public_id": employee.public_id},
            ),
            payload,
        )

        self.assertEqual(response.status_code, 302)
        employee.refresh_from_db()
        self.assertEqual(employee.division, self.second_division)
        self.assertEqual(Employee.objects.filter(personnel_number="PM-001").count(), 1)
        self.assertEqual(employee.contact_profile.operational_phone, "2002")
        self.assertTrue(
            PersonnelChangeRecord.objects.filter(
                employee=employee,
                action=PersonnelChangeAction.UPDATE,
            ).exists()
        )

    def test_edit_right_closes_previous_revision_and_creates_linked_grant(self) -> None:
        employee = self.create_employee()
        previous = EmployeeOperationalRight.objects.create(
            employee=employee,
            right_definition=self.right_definition,
            source_marker="+",
            source_reference="SYNTHETIC-R1",
            source_file_sha256="a" * 64,
            source_row_number=1,
            valid_from=date(2026, 1, 1),
        )

        response = self.client.post(
            reverse(
                "organizations:operational_right_edit",
                kwargs={
                    "employee_public_id": employee.public_id,
                    "record_public_id": previous.public_id,
                },
            ),
            {
                "right_definition": str(self.right_definition.id),
                "source_marker": "+1",
                "qualifier": "После подтверждения начальником смены",
                "scope_text": "Синтетическая оперативная область",
                "valid_from": "2026-02-01",
                "valid_until": "",
                "source_reference": "SYNTHETIC-R2",
                "is_active": "on",
                "change_reason": "Новая редакция матрицы",
            },
        )

        self.assertEqual(response.status_code, 302)
        previous.refresh_from_db()
        self.assertFalse(previous.is_active)
        current = EmployeeOperationalRight.objects.get(
            employee=employee,
            right_definition=self.right_definition,
            is_active=True,
        )
        self.assertEqual(current.source_marker, "+1")
        grant = OperationalAuthorityGrant.objects.get(source_operational_right=current)
        self.assertEqual(grant.basis_status, "VERIFY")
        self.assertTrue(
            PersonnelChangeRecord.objects.filter(
                employee=employee,
                action=PersonnelChangeAction.RIGHT,
            ).exists()
        )

    def test_xlsx_template_preview_and_publish_creates_employee(self) -> None:
        template_bytes = build_personnel_template("INTERNAL_MATRIX")
        workbook = load_workbook(io.BytesIO(template_bytes))
        sheet = workbook.active
        headers = {cell.value: cell.column for cell in sheet[1]}
        sheet.cell(2, headers["Табельный номер"], "BATCH-001")
        sheet.cell(2, headers["Фамилия"], "Пакетный")
        sheet.cell(2, headers["Имя"], "Сотрудник")
        sheet.cell(2, headers["Отчество"], "Тестович")
        sheet.cell(2, headers["Подразделение"], "Оперативная служба")
        sheet.cell(2, headers["Должность"], "Электромонтёр")
        sheet.cell(2, headers["Рабочее место"], "Сменный персонал")
        sheet.cell(2, headers["Категория персонала"], "ОП")
        sheet.cell(2, headers["Группа по электробезопасности"], "IV")
        sheet.cell(2, headers["Класс напряжения"], "до и выше 1000 В")
        sheet.cell(2, headers["Оперативный телефон"], "3003")
        right_header = next(
            value
            for value in headers
            if value.startswith("Право: switching_operation|")
        )
        sheet.cell(2, headers[right_header], "+1 | После подтверждения")
        payload = io.BytesIO()
        workbook.save(payload)
        upload = SimpleUploadedFile(
            "personnel-matrix.xlsx",
            payload.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

        response = self.client.post(
            reverse("organizations:personnel_import_upload"),
            {
                "target_organization": str(self.organization.id),
                "source_organization": "",
                "import_kind": "INTERNAL_MATRIX",
                "source_reference": "SYNTHETIC-BATCH-R1",
                "effective_from": "2026-03-01",
                "workbook": upload,
            },
        )

        self.assertEqual(response.status_code, 302)
        batch = PersonnelImportBatch.objects.get(uploaded_name="personnel-matrix.xlsx")
        self.assertEqual(batch.status, PersonnelImportStatus.PREVIEW)
        self.assertEqual(batch.preview["summary"]["create"], 1)

        publish = self.client.post(
            reverse(
                "organizations:personnel_import_publish",
                kwargs={"public_id": batch.public_id},
            ),
            {"selected_rows": ["2"]},
        )

        self.assertEqual(publish.status_code, 302)
        batch.refresh_from_db()
        self.assertEqual(batch.status, PersonnelImportStatus.PUBLISHED)
        employee = Employee.objects.get(
            organization=self.organization,
            personnel_number="BATCH-001",
        )
        self.assertEqual(employee.contact_profile.operational_phone, "3003")
        self.assertEqual(employee.qualifications.get(is_active=True).electrical_safety_group, "IV")
        right = employee.operational_rights.get(is_active=True)
        self.assertEqual(right.source_marker, "+1")
        self.assertTrue(
            OperationalAuthorityGrant.objects.filter(
                source_operational_right=right,
            ).exists()
        )

    def test_template_download_is_available(self) -> None:
        response = self.client.get(
            reverse(
                "organizations:personnel_import_template",
                kwargs={"import_kind": "EXTERNAL_DIRECTORY"},
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(io.BytesIO(response.content), read_only=True)
        self.assertEqual(workbook.active.title, "Персонал")
