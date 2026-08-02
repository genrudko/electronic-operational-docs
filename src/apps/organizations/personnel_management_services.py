from __future__ import annotations

import hashlib
import io
import re
from datetime import UTC, date, datetime, time, timedelta
from typing import Any

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation

from apps.normatives.evidence import sha256_digest

from .authority_models import (
    AuthorityBasisStatus,
    AuthorityScopeKind,
    OperationalAuthorityGrant,
)
from .models import (
    Division,
    Employee,
    EmployeeOperationalRight,
    EmployeeQualification,
    OperationalRightDefinition,
    Organization,
    Position,
    Workplace,
)
from .personnel_management_models import (
    EmployeeContactProfile,
    ExternalOperationalContact,
    ExternalOperationalRelationKind,
    PersonnelChangeAction,
    PersonnelChangeRecord,
    PersonnelImportBatch,
    PersonnelImportKind,
)


RIGHT_ACTION_CODES = {
    "dispatch_request_submit": "REQUEST.DISPATCH.SUBMIT",
    "dispatch_request_approve": "REQUEST.DISPATCH.APPROVE",
    "operational_request_submit": "REQUEST.OPERATIONAL.SUBMIT",
    "operational_request_approve": "REQUEST.OPERATIONAL.APPROVE",
    "interlock_release": "INTERLOCK.RELEASE",
    "worksite_preparation_admission_authorize": "WORKSITE.AUTHORIZE",
    "work_permit_issue": "WORK.PERMIT.ISSUE",
    "responsible_work_manager": "WORK.RESPONSIBLE_MANAGER",
    "admitting_person": "WORK.ADMIT",
    "work_supervisor": "WORK.SUPERVISE",
    "observer": "WORK.OBSERVE",
    "crew_member": "WORK.CREW_MEMBER",
    "sole_inspection": "EQUIPMENT.INSPECT",
    "operational_communications": "COMMUNICATIONS.OPERATIONAL",
    "switching_operation": "SWITCHING.EXECUTE",
    "switching_control": "SWITCHING.CONTROL",
    "electrical_installation_scope": "ELECTRICAL_INSTALLATION.ACCESS",
    "work_at_height": "SPECIAL_WORK.HEIGHT",
    "live_work": "SPECIAL_WORK.LIVE",
    "induced_voltage_work": "SPECIAL_WORK.INDUCED_VOLTAGE",
    "high_voltage_testing": "SPECIAL_WORK.HIGH_VOLTAGE_TEST",
    "rza_maintenance_category": "RZA.MAINTENANCE",
}

MARKER_CONDITIONS = {
    "+1": "Предоставлено с дополнительным условием 1 из документа-основания.",
    "+2": "Предоставлено с дополнительным условием 2 из документа-основания.",
    "+3": "Предоставлено с дополнительным условием 3 из документа-основания.",
}

RELATION_ALIASES = {
    "диспетчерский": ExternalOperationalRelationKind.DISPATCH,
    "диспетчер": ExternalOperationalRelationKind.DISPATCH,
    "оперативный": ExternalOperationalRelationKind.OPERATIONAL,
    "руководство": ExternalOperationalRelationKind.MANAGEMENT,
    "руководитель": ExternalOperationalRelationKind.MANAGEMENT,
    "цус": ExternalOperationalRelationKind.CONTROL_CENTER,
    "центр управления сетями": ExternalOperationalRelationKind.CONTROL_CENTER,
    "коммерческий": ExternalOperationalRelationKind.COMMERCIAL_DISPATCH,
    "смежный энергообъект": ExternalOperationalRelationKind.RELATED_SITE,
}


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).split())


def _normalized(value: object) -> str:
    return _text(value).casefold().replace("ё", "е")


def _stable_code(prefix: str, *parts: object, length: int = 12) -> str:
    payload = "|".join(_normalized(item) for item in parts)
    digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:length].upper()
    return f"{prefix}-{digest}"


def _date_value(value: object) -> str:
    if isinstance(value, (date, datetime)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return _text(value)


def employee_snapshot(employee: Employee) -> dict[str, Any]:
    contact = EmployeeContactProfile.objects.filter(employee=employee).first()
    return {
        "employee_id": employee.id,
        "public_id": str(employee.public_id),
        "organization_id": employee.organization_id,
        "division_id": employee.division_id,
        "position_id": employee.position_id,
        "workplace_id": employee.workplace_id,
        "personnel_number": employee.personnel_number,
        "full_name": employee.full_name,
        "employment_start": employee.employment_start.isoformat(),
        "employment_end": employee.employment_end.isoformat() if employee.employment_end else None,
        "is_active": employee.is_active,
        "contact": {
            "primary_phone": contact.primary_phone if contact else "",
            "operational_phone": contact.operational_phone if contact else "",
            "email": contact.email if contact else "",
            "availability_schedule": contact.availability_schedule if contact else "",
            "is_round_the_clock": contact.is_round_the_clock if contact else False,
        },
        "qualifications": list(
            employee.qualifications.order_by("valid_from", "id").values(
                "id",
                "personnel_category",
                "electrical_safety_group",
                "voltage_scope",
                "electrical_installation_scope",
                "valid_from",
                "valid_until",
                "is_active",
                "source_reference",
            )
        ),
        "special_qualifications": list(
            employee.special_qualifications.order_by("valid_from", "id").values(
                "id",
                "kind",
                "level",
                "scope_text",
                "valid_from",
                "valid_until",
                "is_active",
                "basis_reference",
            )
        ),
        "rights": list(
            employee.operational_rights.select_related("right_definition")
            .order_by("right_definition__display_order", "valid_from", "id")
            .values(
                "id",
                "right_definition__code",
                "source_marker",
                "qualifier",
                "scope_text",
                "valid_from",
                "valid_until",
                "is_active",
                "source_reference",
            )
        ),
    }


def record_personnel_change(
    *,
    user,
    action: str,
    reason: str,
    employee: Employee | None = None,
    batch: PersonnelImportBatch | None = None,
    before: dict[str, Any] | None = None,
    after: dict[str, Any] | None = None,
) -> PersonnelChangeRecord:
    return PersonnelChangeRecord.objects.create(
        employee=employee,
        batch=batch,
        action=action,
        reason=reason,
        before_snapshot=before or {},
        after_snapshot=after or {},
        changed_by=user,
    )


def manual_source_hash(*, employee: Employee, record_kind: str, basis: str) -> str:
    return sha256_digest(
        {
            "employee_public_id": str(employee.public_id),
            "record_kind": record_kind,
            "basis": basis,
            "recorded_at": timezone.now(),
        }
    )


def _header_map(values: list[object]) -> dict[str, int]:
    return {_normalized(value): index for index, value in enumerate(values) if _text(value)}


def _column(headers: dict[str, int], *aliases: str) -> int | None:
    for alias in aliases:
        normalized = _normalized(alias)
        if normalized in headers:
            return headers[normalized]
    return None


def _cell(values: list[object], index: int | None) -> str:
    if index is None or index >= len(values):
        return ""
    return _text(values[index])


def _split_full_name(value: str) -> tuple[str, str, str]:
    parts = value.split()
    if len(parts) < 2:
        raise ValidationError("ФИО должно содержать как минимум фамилию и имя.")
    return parts[0], parts[1], " ".join(parts[2:])


def _bool_value(value: str) -> bool:
    return _normalized(value) in {"1", "да", "true", "+", "круглосуточно", "24/7"}


def _relation_kind(value: str) -> str:
    normalized = _normalized(value)
    for marker, kind in RELATION_ALIASES.items():
        if marker in normalized:
            return kind
    return ExternalOperationalRelationKind.OPERATIONAL


def _find_employee(
    *,
    organization: Organization,
    personnel_number: str,
    last_name: str,
    first_name: str,
    middle_name: str,
) -> Employee | None:
    if personnel_number:
        matched = Employee.objects.filter(
            organization=organization,
            personnel_number=personnel_number,
        ).first()
        if matched:
            return matched
    return Employee.objects.filter(
        organization=organization,
        last_name__iexact=last_name,
        first_name__iexact=first_name,
        middle_name__iexact=middle_name,
    ).first()


def _parse_right_header(value: object) -> str | None:
    header = _text(value)
    if not _normalized(header).startswith("право:"):
        return None
    token = header.split(":", 1)[1].strip()
    return token.split("|", 1)[0].strip().lower()


def _parse_marker(value: str) -> tuple[str, str]:
    raw = _text(value)
    if not raw or raw in {"-", "—", "0"}:
        return "", ""
    parts = [part.strip() for part in re.split(r"\s*\|\s*", raw, maxsplit=1)]
    marker = parts[0]
    qualifier = parts[1] if len(parts) > 1 else MARKER_CONDITIONS.get(marker, "")
    return marker, qualifier


def parse_personnel_workbook(
    *,
    file_bytes: bytes,
    import_kind: str,
    target_organization: Organization,
    source_organization: Organization | None,
) -> dict[str, Any]:
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError(f"Не удалось открыть XLSX: {exc}") from exc
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValidationError("В файле нет строк.")
    header_index = next(
        (index for index, row in enumerate(rows[:20]) if any(_text(value) for value in row)),
        None,
    )
    if header_index is None:
        raise ValidationError("Не найдена строка заголовков.")
    headers = _header_map(list(rows[header_index]))
    parsed_rows: list[dict[str, Any]] = []
    file_errors: list[str] = []

    if import_kind == PersonnelImportKind.INTERNAL_MATRIX:
        required = {
            "personnel_number": _column(headers, "Табельный номер", "Внутренний номер"),
            "last_name": _column(headers, "Фамилия"),
            "first_name": _column(headers, "Имя"),
            "division": _column(headers, "Подразделение"),
            "position": _column(headers, "Должность"),
        }
        missing = [name for name, index in required.items() if index is None]
        if missing:
            raise ValidationError("Не найдены обязательные колонки: " + ", ".join(missing))
        right_columns = {
            index: code
            for index, value in enumerate(rows[header_index])
            if (code := _parse_right_header(value))
        }
        known_rights = set(
            OperationalRightDefinition.objects.filter(is_active=True).values_list("code", flat=True)
        )
        unknown_rights = sorted(set(right_columns.values()) - known_rights)
        if unknown_rights:
            file_errors.append("Неизвестные коды прав: " + ", ".join(unknown_rights))
        for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            values = list(row)
            if not any(_text(value) for value in values):
                continue
            errors: list[str] = []
            personnel_number = _cell(values, required["personnel_number"])
            last_name = _cell(values, required["last_name"])
            first_name = _cell(values, required["first_name"])
            middle_name = _cell(values, _column(headers, "Отчество"))
            division = _cell(values, required["division"])
            position = _cell(values, required["position"])
            if not all((personnel_number, last_name, first_name, division, position)):
                errors.append("Не заполнены обязательные сведения о сотруднике.")
            matched = _find_employee(
                organization=target_organization,
                personnel_number=personnel_number,
                last_name=last_name,
                first_name=first_name,
                middle_name=middle_name,
            ) if not errors else None
            rights = []
            for index, code in right_columns.items():
                marker, qualifier = _parse_marker(_cell(values, index))
                if marker and code in known_rights:
                    rights.append({"code": code, "marker": marker, "qualifier": qualifier})
            parsed_rows.append(
                {
                    "row_number": row_number,
                    "selected": not errors,
                    "action": "UPDATE" if matched else "CREATE",
                    "matched_employee_id": matched.id if matched else None,
                    "errors": errors,
                    "warnings": [],
                    "employee": {
                        "personnel_number": personnel_number,
                        "last_name": last_name,
                        "first_name": first_name,
                        "middle_name": middle_name,
                        "division": division,
                        "position": position,
                        "workplace": _cell(values, _column(headers, "Рабочее место")),
                    },
                    "contact": {
                        "primary_phone": _cell(values, _column(headers, "Основной телефон", "Телефон")),
                        "operational_phone": _cell(values, _column(headers, "Оперативный телефон")),
                        "email": _cell(values, _column(headers, "Электронная почта", "Email")),
                        "availability_schedule": _cell(values, _column(headers, "Часы работы", "График")),
                        "is_round_the_clock": _bool_value(_cell(values, _column(headers, "Круглосуточно", "24/7"))),
                    },
                    "qualification": {
                        "personnel_category": _cell(values, _column(headers, "Категория персонала")),
                        "electrical_safety_group": _cell(values, _column(headers, "Группа по электробезопасности")),
                        "voltage_scope": _cell(values, _column(headers, "Класс напряжения")),
                        "electrical_installation_scope": _cell(values, _column(headers, "Область электроустановок")),
                    },
                    "rights": rights,
                }
            )
    else:
        full_name_index = _column(headers, "ФИО", "Ф.И.О.", "Ф.И.О")
        position_index = _column(headers, "Должность")
        division_index = _column(headers, "Подразделение", "Подразделение / объект", "Энергообъект")
        if full_name_index is None or position_index is None or division_index is None:
            raise ValidationError("Для внешнего списка требуются ФИО, Должность и Подразделение / объект.")
        if source_organization is None:
            raise ValidationError("Не указана организация-источник внешнего списка.")
        for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            values = list(row)
            if not any(_text(value) for value in values):
                continue
            errors = []
            full_name = _cell(values, full_name_index)
            try:
                last_name, first_name, middle_name = _split_full_name(full_name)
            except ValidationError as exc:
                last_name = first_name = middle_name = ""
                errors.append(str(exc))
            position = _cell(values, position_index)
            division = _cell(values, division_index)
            personnel_number = _cell(values, _column(headers, "Табельный номер", "Внутренний номер"))
            if not personnel_number and full_name:
                personnel_number = _stable_code("EXT", source_organization.code, full_name)
            if not position or not division:
                errors.append("Не заполнены должность или подразделение / объект.")
            matched = _find_employee(
                organization=source_organization,
                personnel_number=personnel_number,
                last_name=last_name,
                first_name=first_name,
                middle_name=middle_name,
            ) if not errors else None
            parsed_rows.append(
                {
                    "row_number": row_number,
                    "selected": not errors,
                    "action": "UPDATE" if matched else "CREATE",
                    "matched_employee_id": matched.id if matched else None,
                    "errors": errors,
                    "warnings": [],
                    "employee": {
                        "personnel_number": personnel_number,
                        "last_name": last_name,
                        "first_name": first_name,
                        "middle_name": middle_name,
                        "division": division,
                        "position": position,
                        "workplace": _cell(values, _column(headers, "Рабочее место")),
                    },
                    "contact": {
                        "primary_phone": _cell(values, _column(headers, "Основной телефон", "Телефон")),
                        "operational_phone": _cell(values, _column(headers, "Оперативный телефон")),
                        "email": _cell(values, _column(headers, "Электронная почта", "Email")),
                        "availability_schedule": _cell(values, _column(headers, "Часы работы", "График")),
                        "is_round_the_clock": _bool_value(_cell(values, _column(headers, "Круглосуточно", "24/7"))),
                    },
                    "external": {
                        "relation_kind": _relation_kind(_cell(values, _column(headers, "Тип персонала", "Роль во взаимодействии"))),
                        "operational_scope": _cell(values, _column(headers, "Область взаимодействия", "Область")),
                        "authority_summary": _cell(values, _column(headers, "Полномочия", "Права")),
                    },
                }
            )

    return {
        "sheet_name": sheet.title,
        "header_row": header_index + 1,
        "rows": parsed_rows,
        "file_errors": file_errors,
        "summary": {
            "total": len(parsed_rows),
            "create": sum(item["action"] == "CREATE" and not item["errors"] for item in parsed_rows),
            "update": sum(item["action"] == "UPDATE" and not item["errors"] for item in parsed_rows),
            "errors": sum(bool(item["errors"]) for item in parsed_rows),
        },
    }


def _get_or_create_division(organization: Organization, name: str) -> Division:
    matched = Division.objects.filter(organization=organization, name__iexact=name).first()
    if matched:
        return matched
    return Division.objects.create(
        organization=organization,
        code=_stable_code("DIV", organization.code, name),
        name=name,
    )


def _get_or_create_position(organization: Organization, name: str) -> Position:
    matched = Position.objects.filter(organization=organization, name__iexact=name).first()
    if matched:
        return matched
    return Position.objects.create(
        organization=organization,
        code=_stable_code("POS", organization.code, name),
        name=name,
        is_operational=any(
            marker in _normalized(name)
            for marker in ("диспетчер", "оператив", "дежурн", "начальник смены")
        ),
    )


def _get_or_create_workplace(
    organization: Organization,
    division: Division,
    name: str,
) -> Workplace | None:
    if not name:
        return None
    matched = Workplace.objects.filter(organization=organization, name__iexact=name).first()
    if matched:
        return matched
    return Workplace.objects.create(
        organization=organization,
        division=division,
        code=_stable_code("WP", organization.code, name),
        name=name,
    )


def _upsert_employee(
    *,
    organization: Organization,
    row: dict[str, Any],
    effective_from: date,
) -> tuple[Employee, dict[str, Any]]:
    data = row["employee"]
    employee = None
    if row.get("matched_employee_id"):
        employee = Employee.objects.filter(
            pk=row["matched_employee_id"],
            organization=organization,
        ).first()
    if employee is None:
        employee = _find_employee(
            organization=organization,
            personnel_number=data["personnel_number"],
            last_name=data["last_name"],
            first_name=data["first_name"],
            middle_name=data["middle_name"],
        )
    before = employee_snapshot(employee) if employee else {}
    division = _get_or_create_division(organization, data["division"])
    position = _get_or_create_position(organization, data["position"])
    workplace = _get_or_create_workplace(organization, division, data.get("workplace", ""))
    if employee is None:
        employee = Employee(
            organization=organization,
            personnel_number=data["personnel_number"],
            employment_start=effective_from,
        )
    employee.division = division
    employee.position = position
    employee.workplace = workplace
    employee.last_name = data["last_name"]
    employee.first_name = data["first_name"]
    employee.middle_name = data["middle_name"]
    employee.is_active = True
    employee.full_clean()
    employee.save()
    return employee, before


def _upsert_contact(employee: Employee, data: dict[str, Any]) -> None:
    contact, _ = EmployeeContactProfile.objects.get_or_create(employee=employee)
    for field in (
        "primary_phone",
        "operational_phone",
        "email",
        "availability_schedule",
        "is_round_the_clock",
    ):
        setattr(contact, field, data.get(field, "" if field != "is_round_the_clock" else False))
    contact.save()


def _publish_qualification(
    *,
    employee: Employee,
    data: dict[str, Any],
    batch: PersonnelImportBatch,
    row_number: int,
) -> None:
    if not any(data.values()):
        return
    EmployeeQualification.objects.filter(employee=employee, is_active=True).update(is_active=False)
    EmployeeQualification.objects.update_or_create(
        employee=employee,
        source_file_sha256=batch.file_sha256,
        source_row_number=row_number,
        defaults={
            "personnel_category": data.get("personnel_category", ""),
            "electrical_safety_group": data.get("electrical_safety_group", ""),
            "voltage_scope": data.get("voltage_scope", ""),
            "electrical_installation_scope": data.get("electrical_installation_scope", ""),
            "valid_from": batch.effective_from,
            "valid_until": None,
            "is_active": True,
            "source_reference": batch.source_reference,
        },
    )


def _publish_right(
    *,
    employee: Employee,
    item: dict[str, str],
    batch: PersonnelImportBatch,
    row_number: int,
) -> None:
    definition = OperationalRightDefinition.objects.get(code=item["code"], is_active=True)
    existing = EmployeeOperationalRight.objects.filter(
        employee=employee,
        right_definition=definition,
        is_active=True,
    ).exclude(source_file_sha256=batch.file_sha256, source_row_number=row_number)
    day_before = batch.effective_from - timedelta(days=1)
    for previous in existing:
        previous.is_active = False
        if previous.valid_until is None or previous.valid_until >= batch.effective_from:
            previous.valid_until = max(previous.valid_from, day_before)
        previous.save()
    source_right, _ = EmployeeOperationalRight.objects.update_or_create(
        employee=employee,
        right_definition=definition,
        source_file_sha256=batch.file_sha256,
        source_row_number=row_number,
        defaults={
            "qualifier": item.get("qualifier", ""),
            "scope_text": employee.organization.name,
            "source_marker": item["marker"],
            "source_reference": batch.source_reference,
            "valid_from": batch.effective_from,
            "valid_until": None,
            "is_active": True,
        },
    )
    start = datetime.combine(batch.effective_from, time.min, tzinfo=UTC)
    status = (
        AuthorityBasisStatus.CONFIRMED
        if item["marker"] == "+"
        else AuthorityBasisStatus.VERIFY
    )
    action_code = RIGHT_ACTION_CODES.get(
        definition.code,
        f"PERSONNEL.RIGHT.{definition.code.upper()}",
    )
    OperationalAuthorityGrant.objects.update_or_create(
        employee=employee,
        action_code=action_code,
        scope_kind=AuthorityScopeKind.ORGANIZATION,
        scope_reference=str(employee.organization_id),
        valid_from=start,
        basis_reference=batch.source_reference,
        defaults={
            "organization": employee.organization,
            "right_definition": definition,
            "scope_label": employee.organization.name,
            "granting_organization": employee.organization,
            "basis_status": status,
            "source_ids": [
                f"PERSONNEL-IMPORT-{batch.public_id}",
                definition.code,
            ],
            "source_operational_right": source_right,
            "valid_until": None,
            "is_active": True,
            "allow_substitution": False,
        },
    )


@transaction.atomic
def publish_import_batch(
    *,
    batch: PersonnelImportBatch,
    selected_rows: set[int],
    user,
) -> dict[str, int]:
    if batch.status != "PREVIEW":
        raise ValidationError("Пакет уже обработан.")
    counters = {"created": 0, "updated": 0, "skipped": 0}
    organization = (
        batch.target_organization
        if batch.import_kind == PersonnelImportKind.INTERNAL_MATRIX
        else batch.source_organization
    )
    if organization is None:
        raise ValidationError("Для публикации не определена организация персонала.")
    for row in batch.preview.get("rows", []):
        if row["row_number"] not in selected_rows or row.get("errors"):
            counters["skipped"] += 1
            continue
        employee, before = _upsert_employee(
            organization=organization,
            row=row,
            effective_from=batch.effective_from,
        )
        _upsert_contact(employee, row.get("contact", {}))
        if batch.import_kind == PersonnelImportKind.INTERNAL_MATRIX:
            _publish_qualification(
                employee=employee,
                data=row.get("qualification", {}),
                batch=batch,
                row_number=row["row_number"],
            )
            for right in row.get("rights", []):
                _publish_right(
                    employee=employee,
                    item=right,
                    batch=batch,
                    row_number=row["row_number"],
                )
        else:
            external = row.get("external", {})
            ExternalOperationalContact.objects.update_or_create(
                employee=employee,
                host_organization=batch.target_organization,
                relation_kind=external.get(
                    "relation_kind",
                    ExternalOperationalRelationKind.OPERATIONAL,
                ),
                valid_from=batch.effective_from,
                defaults={
                    "operational_scope": external.get("operational_scope", ""),
                    "authority_summary": external.get("authority_summary", ""),
                    "valid_until": None,
                    "basis_reference": batch.source_reference,
                    "is_active": True,
                },
            )
        after = employee_snapshot(employee)
        action = (
            PersonnelChangeAction.UPDATE
            if before
            else PersonnelChangeAction.CREATE
        )
        record_personnel_change(
            user=user,
            employee=employee,
            batch=batch,
            action=action,
            reason=f"Публикация XLSX: {batch.source_reference}",
            before=before,
            after=after,
        )
        counters["updated" if before else "created"] += 1
    batch.mark_published(user)
    record_personnel_change(
        user=user,
        batch=batch,
        action=PersonnelChangeAction.IMPORT_PUBLISH,
        reason=f"Публикация пакета {batch.uploaded_name}",
        after=counters,
    )
    return counters


def create_import_batch(*, form, user) -> PersonnelImportBatch:
    workbook = form.cleaned_data["workbook"]
    file_bytes = workbook.read()
    file_hash = hashlib.sha256(file_bytes).hexdigest()
    if PersonnelImportBatch.objects.filter(file_sha256=file_hash).exists():
        raise ValidationError("Этот файл уже загружался. Откройте существующий пакет импорта.")
    preview = parse_personnel_workbook(
        file_bytes=file_bytes,
        import_kind=form.cleaned_data["import_kind"],
        target_organization=form.cleaned_data["target_organization"],
        source_organization=form.cleaned_data.get("source_organization"),
    )
    batch = PersonnelImportBatch.objects.create(
        target_organization=form.cleaned_data["target_organization"],
        source_organization=form.cleaned_data.get("source_organization"),
        import_kind=form.cleaned_data["import_kind"],
        uploaded_name=workbook.name,
        file_sha256=file_hash,
        sheet_name=preview["sheet_name"],
        source_reference=form.cleaned_data["source_reference"],
        effective_from=form.cleaned_data["effective_from"],
        preview=preview,
        validation_errors=preview["file_errors"],
        uploaded_by=user,
    )
    record_personnel_change(
        user=user,
        batch=batch,
        action=PersonnelChangeAction.IMPORT_PREVIEW,
        reason=f"Предварительный просмотр {workbook.name}",
        after=preview["summary"],
    )
    return batch


def build_personnel_template(import_kind: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Персонал"
    if import_kind == PersonnelImportKind.INTERNAL_MATRIX:
        rights = list(
            OperationalRightDefinition.objects.filter(is_active=True).order_by(
                "display_order",
                "name",
            )
        )
        headers = [
            "Табельный номер",
            "Фамилия",
            "Имя",
            "Отчество",
            "Подразделение",
            "Должность",
            "Рабочее место",
            "Категория персонала",
            "Группа по электробезопасности",
            "Класс напряжения",
            "Область электроустановок",
            "Основной телефон",
            "Оперативный телефон",
            "Электронная почта",
            "Часы работы",
            "Круглосуточно",
        ] + [f"Право: {right.code}|{right.name}" for right in rights]
    else:
        headers = [
            "ФИО",
            "Табельный номер",
            "Подразделение / объект",
            "Должность",
            "Тип персонала",
            "Основной телефон",
            "Оперативный телефон",
            "Электронная почта",
            "Часы работы",
            "Круглосуточно",
            "Область взаимодействия",
            "Полномочия",
        ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[cell_column(index)].width = min(max(len(header) * 0.9, 16), 42)
        if header.startswith("Право:"):
            sheet.cell(1, index).comment = Comment(
                "Допустимые значения: +, -, +1, +2, +3. После символа | можно указать условие.",
                "ЭОД",
            )
    if import_kind == PersonnelImportKind.INTERNAL_MATRIX:
        category_validation = DataValidation(
            type="list",
            formula1='"АТП,ОП,АТП/ОП,ОРП,РП"',
            allow_blank=True,
        )
        group_validation = DataValidation(
            type="list",
            formula1='"II,III,IV,V"',
            allow_blank=True,
        )
        marker_validation = DataValidation(
            type="list",
            formula1='"+,-,+1,+2,+3"',
            allow_blank=True,
        )
        sheet.add_data_validation(category_validation)
        sheet.add_data_validation(group_validation)
        sheet.add_data_validation(marker_validation)
        category_validation.add(f"H2:H1000")
        group_validation.add(f"I2:I1000")
        if len(headers) > 16:
            marker_validation.add(f"Q2:{cell_column(len(headers))}1000")
    else:
        relation_validation = DataValidation(
            type="list",
            formula1=(
                '"Диспетчерский персонал,Оперативный персонал,Руководство,'
                'Персонал ЦУС,Коммерческий диспетчер,Смежный энергообъект"'
            ),
            allow_blank=False,
        )
        sheet.add_data_validation(relation_validation)
        relation_validation.add("E2:E1000")
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def cell_column(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result
